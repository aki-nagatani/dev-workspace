#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
人事の目標管理シート（Excel）を読み取り、JSON またはテキストで stdout に出すだけのツール。

- mokuhyo-excel-to-markdown SKILL のセル正本（BH4・BF6・枝番基準行・列オフセット）に従う。
- **Markdown への書き込みは行わない**（myrules の「スクリプトで .md を生成・上書き禁止」と整合）。

使用例:
  python scripts/read_hr_goal_sheet_xls.py "D:/path/72858（長谷晃英）目標-53上.xls"
  python scripts/read_hr_goal_sheet_xls.py --format text file1.xls file2.xls
  python scripts/read_hr_goal_sheet_xls.py --sheet 目標管理シート path.xls

依存:
  - .xls: xlrd（pip install xlrd）
  - .xlsx / .xlsm: openpyxl（pip install openpyxl）
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def column_letters_to_index(letters: str) -> int:
    """A1 列記法（例: C, BH）を 0 始まり列インデックスに変換する。"""
    n = 0
    for c in letters.upper().strip():
        if not ("A" <= c <= "Z"):
            msg = f"Invalid column letters: {letters!r}"
            raise ValueError(msg)
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    s = str(value).strip()
    return s


def _read_cell_xlrd(sheet: Any, row_1based: int, col_letters: str) -> str:
    import xlrd  # noqa: PLC0415

    rowx = row_1based - 1
    colx = column_letters_to_index(col_letters)
    try:
        v = sheet.cell_value(rowx, colx)
    except IndexError:
        return ""
    if sheet.cell_type(rowx, colx) == xlrd.XL_CELL_EMPTY:
        return ""
    return _cell_to_str(v)


def _read_cell_openpyxl(ws: Any, row_1based: int, col_letters: str) -> str:
    # 公開パッケージ境界は openpyxl.utils（utils.cell は内部モジュールのため解析器が解決しにくい）
    from openpyxl.utils import column_index_from_string  # noqa: PLC0415

    col_1based = column_index_from_string(col_letters)
    v = ws.cell(row=row_1based, column=col_1based).value
    return _cell_to_str(v)


def _load_sheet_xlrd(path: Path, sheet_name: str | None):
    import xlrd  # noqa: PLC0415

    book = xlrd.open_workbook(str(path), formatting_info=False)
    if sheet_name:
        try:
            sh = book.sheet_by_name(sheet_name)
        except xlrd.XLRDError as e:
            names = book.sheet_names()
            msg = f"シート '{sheet_name}' が見つかりません。利用可能: {names}"
            raise SystemExit(msg) from e
    else:
        sh = book.sheet_by_index(0)
    return sh, "xlrd"


def _load_sheet_openpyxl(path: Path, sheet_name: str | None):
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            msg = f"シート '{sheet_name}' が見つかりません。利用可能: {wb.sheetnames}"
            raise SystemExit(msg)
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    return ws, "openpyxl"


def _staff_number_six_digits(raw: str) -> str:
    """BF6 の値を Markdown 用に 6 桁ゼロ埋め（SKILL「社員番号（BF6）と Markdown」）。"""
    s = raw.strip()
    if not s:
        return ""
    try:
        n = int(float(s))
    except ValueError:
        return s
    if n < 0 or n > 999_999:
        return s
    return f"{n:06d}"


def _branch_base_row_excel(k: int) -> int:
    """枝番 k（1〜5）の基準行（Excel 1 始まり）。SKILL: r = 18 + 4*(k-1)。"""
    if k < 1 or k > 5:
        msg = f"枝番は 1〜5: {k}"
        raise ValueError(msg)
    return 18 + 4 * (k - 1)


def extract_header(read_cell) -> dict[str, Any]:
    """BH4, BS4, BF6, BI1, BM1。"""
    raw_bf6 = read_cell(6, "BF")
    return {
        "name_bh4": read_cell(4, "BH"),
        "grade_bs4": read_cell(4, "BS"),
        "staff_bf6_raw": raw_bf6,
        "staff_number_6digits": _staff_number_six_digits(raw_bf6),
        "year_bi1": read_cell(1, "BI"),
        "half_bm1": read_cell(1, "BM"),
    }


def extract_branch(read_cell, k: int) -> dict[str, Any]:
    """枝番 k の SKILL 表に沿ったセル。"""
    r = _branch_base_row_excel(k)
    return {
        "k": k,
        "base_row_excel": r,
        "branch_label_a": read_cell(r, "A"),
        "goal_full_period_c": read_cell(r, "C"),
        "target_full_u": read_cell(r, "U"),
        "minimum_full_u": read_cell(r + 1, "U"),
        "goal_half_c": read_cell(r + 2, "C"),
        "target_half_u": read_cell(r + 2, "U"),
        "minimum_half_u": read_cell(r + 3, "U"),
        "challenge_ac": read_cell(r, "AC"),
        "weight_af": read_cell(r, "AF"),
        "reason_ai": read_cell(r, "AI"),
        "comment_as": read_cell(r, "AS"),
        "achieve_self_bd": read_cell(r, "BD"),
        "achieve_eval_bs": read_cell(r, "BS"),
    }


def _period_from_year(year_str: str) -> str | None:
    """SKILL: 期番号 = 西暦年 - 1971。数値化できなければ None。"""
    s = year_str.strip()
    if not s:
        return None
    try:
        # Excel や xlrd が float で返す年を想定
        y = float(s)
        yi = int(y) if y == int(y) else int(y)
        return str(yi - 1971)
    except ValueError:
        return None


def read_file(path: Path, sheet_name: str | None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        sh, backend = _load_sheet_xlrd(path, sheet_name)

        def read_cell(row_1based: int, col_letters: str) -> str:
            return _read_cell_xlrd(sh, row_1based, col_letters)

    elif suffix in (".xlsx", ".xlsm"):
        ws, backend = _load_sheet_openpyxl(path, sheet_name)

        def read_cell(row_1based: int, col_letters: str) -> str:
            return _read_cell_openpyxl(ws, row_1based, col_letters)

    else:
        msg = f"未対応の拡張子: {suffix}（.xls / .xlsx / .xlsm のみ）"
        raise SystemExit(msg)

    header = extract_header(read_cell)
    year_raw = str(header.get("year_bi1", ""))
    period = _period_from_year(year_raw)

    branches = [extract_branch(read_cell, k) for k in range(1, 6)]

    return {
        "path": str(path.resolve()),
        "sheet": sheet_name or "(既定: 先頭シート)",
        "backend": backend,
        "header": header,
        "computed": {
            "period_from_bi1_rule": period,
            "note": "期番号は BI1 の西暦から 西暦-1971。Markdown の ## 期 と突合せは別途。",
        },
        "branches": branches,
    }


def _preview_line(text: str, max_len: int = 80) -> str:
    one = text.replace("\n", " ").strip()
    if len(one) <= max_len:
        return one
    return one[:max_len] + "…"


def _format_text(data: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"ファイル: {data['path']}")
    lines.append(f"シート: {data['sheet']}")
    lines.append(f"読み取り: {data['backend']}")
    h = data["header"]
    lines.append(f"BH4 名前: {h['name_bh4']}")
    lines.append(f"BS4 等級: {h['grade_bs4']}")
    lines.append(f"BF6 社員番号（生）: {h['staff_bf6_raw']}")
    lines.append(f"Markdown 用（6桁）: {h['staff_number_6digits']}")
    lines.append(f"BI1 年度: {h['year_bi1']}")
    lines.append(f"BM1 上/下期: {h['half_bm1']}")
    cp = data["computed"].get("period_from_bi1_rule")
    if cp:
        lines.append(f"計算期（BI1-1971 ルール）: {cp}期")
    lines.append("")
    for b in data["branches"]:
        lines.append(f"--- 枝 {b['k']} （基準行 Excel {b['base_row_excel']}） A={b['branch_label_a']!r}")
        lines.append(f"  通期目標 C{b['base_row_excel']}: {_preview_line(b['goal_full_period_c'])}")
        lines.append(
            f"  上期/下期目標 C{b['base_row_excel'] + 2}: {_preview_line(b['goal_half_c'])}",
        )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="目標管理シート Excel を読み取り stdout のみに出力（.md へは書かない）。",
    )
    parser.add_argument(
        "paths",
        nargs="+",
        type=Path,
        help="目標管理シートの .xls / .xlsx / .xlsm パス",
    )
    parser.add_argument(
        "--sheet",
        default="目標管理シート",
        help="シート名（既定: 目標管理シート。--sheet \"\" で先頭シート）",
    )
    parser.add_argument(
        "--format",
        choices=("json", "text"),
        default="json",
        help="出力形式（既定: json）",
    )
    args = parser.parse_args()
    sheet_arg: str | None = args.sheet if args.sheet != "" else None

    results: list[dict[str, Any]] = []
    for p in args.paths:
        if not p.is_file():
            print(f"ファイルが見つかりません: {p}", file=sys.stderr)
            sys.exit(1)
        results.append(read_file(p, sheet_arg))

    if args.format == "json":
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        for i, data in enumerate(results):
            if i:
                print("\n" + "=" * 60 + "\n")
            print(_format_text(data))


if __name__ == "__main__":
    main()
