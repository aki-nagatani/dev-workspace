"""案件向け「課題管理表」Excel 雛形を生成する。

用途:
  - 担当案件の課題・タスク棚卸（週次更新・定例での確認）
  - 未クローズ課題の期日・担当者・対応方針の見える化

出力先は引数で指定（既定は stdout にパスのみ表示）。
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "No.",
    "登録日",
    "案件",
    "フェーズ",
    "課題タイトル",
    "課題内容",
    "区分",
    "優先度",
    "状態",
    "担当者",
    "期日",
    "対応方針",
    "進捗・対応履歴",
    "起票者",
    "完了日",
    "最終更新日",
    "備考",
]

COL_WIDTHS = {
    "A": 6,
    "B": 12,
    "C": 14,
    "D": 12,
    "E": 28,
    "F": 36,
    "G": 12,
    "H": 8,
    "I": 10,
    "J": 12,
    "K": 12,
    "L": 28,
    "M": 32,
    "N": 12,
    "O": 12,
    "P": 12,
    "Q": 20,
}

STATUS_VALUES = "未着手,対応中,確認待ち,保留,完了"
PRIORITY_VALUES = "高,中,低"
CATEGORY_VALUES = "障害,仕様確認,改善,タスク,問合せ,その他"
PHASE_VALUES = "要件,設計,製造,単体,結合,総合,移行,運用,その他"

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = THIN
    ws.row_dimensions[1].height = 32


def _apply_data_validations(ws, max_row: int = 500) -> None:
    ranges = {
        "G": CATEGORY_VALUES,
        "H": PRIORITY_VALUES,
        "I": STATUS_VALUES,
        "D": PHASE_VALUES,
    }
    for col, values in ranges.items():
        dv = DataValidation(
            type="list",
            formula1=f'"{values}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "リストから選択してください"
        dv.errorTitle = "入力エラー"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")


def _add_sample_rows(ws) -> None:
    samples = [
        [
            1,
            "2026-08-01",
            "申込書",
            "結合",
            "（記入例）進捗報告の定量値化が未定着",
            "顧客向け進捗報告に定性記述が残り、定量指標の運用が不安定",
            "改善",
            "中",
            "対応中",
            "桐山",
            "2026-08-15",
            "報告フォーマットを固定し、週次で数値欄を必須化",
            "2026-08-05 テンプレ案をレビュー中",
            "長谷",
            "",
            "2026-08-05",
            "記入例。運用開始時に削除可",
        ],
        [
            2,
            "2026-08-03",
            "申込書",
            "設計",
            "（記入例）テスト仕様書共通フォーマットの顧客合意",
            "共通フォーマット案はあるが顧客合意・運用開始前",
            "タスク",
            "高",
            "未着手",
            "桐山",
            "2026-08-29",
            "顧客GL定例で合意取得し、次案件から適用",
            "",
            "長谷",
            "",
            "2026-08-03",
            "記入例。運用開始時に削除可",
        ],
    ]
    body_font = Font(name="Yu Gothic", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    for r_idx, row in enumerate(samples, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = body_font
            cell.alignment = wrap
            cell.border = THIN
            if c_idx in (2, 11, 15, 16) and value:
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r_idx].height = 48

    # 空行の体裁（入力用）
    for r_idx in range(4, 52):
        for c_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(r_idx, c_idx, "")
            cell.font = body_font
            cell.alignment = wrap
            cell.border = THIN
        ws.cell(r_idx, 1, r_idx - 1)


def _add_conditional_formatting(ws, max_row: int = 500) -> None:
    # 完了行を薄く
    done_fill = PatternFill("solid", fgColor="E2EFDA")
    ws.conditional_formatting.add(
        f"A2:Q{max_row}",
        FormulaRule(formula=['$I2="完了"'], fill=done_fill),
    )
    # 期日超過かつ未完了
    overdue_fill = PatternFill("solid", fgColor="FCE4D6")
    ws.conditional_formatting.add(
        f"A2:Q{max_row}",
        FormulaRule(
            formula=['AND($K2<>"",$K2<TODAY(),$I2<>"完了")'],
            fill=overdue_fill,
        ),
    )
    # 優先度 高
    high_font = Font(name="Yu Gothic", color="C00000", bold=True)
    ws.conditional_formatting.add(
        f"H2:H{max_row}",
        FormulaRule(formula=['$H2="高"'], font=high_font),
    )


def _build_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("サマリ", 0)
    title_font = Font(name="Yu Gothic", size=14, bold=True, color="1F4E79")
    label_font = Font(name="Yu Gothic", size=11, bold=True)
    value_font = Font(name="Yu Gothic", size=11)
    tip_font = Font(name="Yu Gothic", size=10, color="666666")

    ws["A1"] = "課題管理表（サマリ）"
    ws["A1"].font = title_font
    ws["A2"] = "週1回以上更新し、チーム／顧客定例で未クローズ課題を棚卸する想定の雛形です。"
    ws["A2"].font = tip_font

    labels = [
        ("A4", "全件数"),
        ("A5", "未クローズ件数"),
        ("A6", "期日超過（未完了）"),
        ("A7", "優先度 高（未完了）"),
        ("A8", "完了件数"),
        ("A9", "期日内解決率（完了のうち）"),
    ]
    for cell_ref, label in labels:
        ws[cell_ref] = label
        ws[cell_ref].font = label_font

    # 課題一覧シート参照（シート名固定）
    ws["B4"] = '=COUNTA(\'課題一覧\'!E2:E500)-COUNTBLANK(\'課題一覧\'!E2:E500)'
    ws["B5"] = '=COUNTIFS(\'課題一覧\'!I2:I500,"<>完了",\'課題一覧\'!E2:E500,"<>")'
    ws["B6"] = (
        '=COUNTIFS(\'課題一覧\'!K2:K500,"<"&TODAY(),'
        '\'課題一覧\'!I2:I500,"<>完了",\'課題一覧\'!E2:E500,"<>")'
    )
    ws["B7"] = (
        '=COUNTIFS(\'課題一覧\'!H2:H500,"高",'
        '\'課題一覧\'!I2:I500,"<>完了",\'課題一覧\'!E2:E500,"<>")'
    )
    ws["B8"] = '=COUNTIF(\'課題一覧\'!I2:I500,"完了")'
    ws["B9"] = (
        '=IF(B8=0,"-",'
        'COUNTIFS(\'課題一覧\'!I2:I500,"完了",\'課題一覧\'!O2:O500,"<>",'
        '\'課題一覧\'!K2:K500,"<>",\'課題一覧\'!O2:O500,"<="&\'課題一覧\'!K2:K500)'
        ")"
    )
    # B9 の配列風は Excel バージョン差があるので、簡易の完了件数比率表示に変更
    ws["B9"] = '=IF(B4=0,"-",TEXT(B8/B4,"0.0%"))'
    ws["C9"] = "（完了件数÷全件数。期日内率は進捗列・完了日で個別確認）"
    ws["C9"].font = tip_font

    for r in range(4, 10):
        ws.cell(r, 2).font = value_font

    ws["A11"] = "運用ルール（推奨）"
    ws["A11"].font = label_font
    rules = [
        "1. 発生したら当日〜翌営業日までに起票（タイトル必須）",
        "2. 未クローズ課題は必ず 担当者・期日・対応方針 を埋める",
        "3. 週1回以上更新し、定例（チーム／顧客）で棚卸する",
        "4. 完了時は状態=完了、完了日を記入。解決済みの詳細は備考または履歴へ要約",
        "5. No. は連番。削除で欠番が出ても再利用しない（参照安定のため）",
        "6. 本ファイルは案件用。組織横断の継続課題は Obsidian Work/課題.md を正本とする",
    ]
    for i, text in enumerate(rules, start=12):
        ws[f"A{i}"] = text
        ws[f"A{i}"].font = value_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 56


def _build_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("使い方")
    font = Font(name="Yu Gothic", size=10)
    bold = Font(name="Yu Gothic", size=11, bold=True, color="1F4E79")
    ws["A1"] = "列の説明"
    ws["A1"].font = bold

    rows = [
        ("No.", "課題の識別番号（連番）"),
        ("登録日", "起票日（yyyy-mm-dd）"),
        ("案件", "申込書／代理店など案件名"),
        ("フェーズ", "発生・対応中の工程"),
        ("課題タイトル", "短い名詞句（一覧で識別できる粒度）"),
        ("課題内容", "事実・影響・背景"),
        ("区分", "障害／仕様確認／改善／タスク／問合せ／その他"),
        ("優先度", "高／中／低"),
        ("状態", "未着手／対応中／確認待ち／保留／完了"),
        ("担当者", "対応の主担当"),
        ("期日", "対応完了の目標日。未クローズは必須"),
        ("対応方針", "何をするか（打ち手）"),
        ("進捗・対応履歴", "日付付きで追記（古い順 or 新しい順で統一）"),
        ("起票者", "登録した人"),
        ("完了日", "クローズした日"),
        ("最終更新日", "行を直した日"),
        ("備考", "顧客共有上の注意、関連票番号など"),
    ]
    ws["A2"] = "列名"
    ws["B2"] = "説明"
    ws["A2"].font = bold
    ws["B2"].font = bold
    for i, (name, desc) in enumerate(rows, start=3):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = desc
        ws[f"A{i}"].font = font
        ws[f"B{i}"].font = font

    ws["A21"] = "フィルタの使い方"
    ws["A21"].font = bold
    ws["A22"] = "課題一覧の見出し行で Autofilter を使用。状態≠完了 で未クローズだけ表示できます。"
    ws["A22"].font = font

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 64


def build_workbook() -> Workbook:
    wb = Workbook()
    # いったんデフォルトシートを課題一覧に
    ws = wb.active
    ws.title = "課題一覧"

    _style_header(ws)
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    _add_sample_rows(ws)
    _apply_data_validations(ws)
    _add_conditional_formatting(ws)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}500"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    _build_summary_sheet(wb)
    _build_guide_sheet(wb)
    # シート順: サマリ / 課題一覧 / 使い方
    # create_sheet("サマリ", 0) 後は [サマリ, 課題一覧]。使い方は末尾追加。
    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="課題管理表 Excel 雛形を生成")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        required=True,
        help="出力 .xlsx パス",
    )
    args = parser.parse_args()
    out: Path = args.output
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out)
    print(str(out.resolve()))


if __name__ == "__main__":
    main()
